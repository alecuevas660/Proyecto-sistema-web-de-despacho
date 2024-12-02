from django.shortcuts import render
from django.views.generic import ListView
from django.db.models import Q
from apps.inventario.models import OrdenDespacho
from openpyxl import Workbook
<<<<<<< HEAD
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from django.http import HttpResponse
from django.utils import timezone
from datetime import datetime
from xhtml2pdf import pisa
from django.template.loader import render_to_string
from openpyxl.utils.exceptions import InvalidFileException
import random

=======
from django.http import HttpResponse
from django.utils import timezone
from xhtml2pdf import pisa
from django.template.loader import render_to_string
>>>>>>> a78dd5d0da9c93d7f98d1c059e1ae4fdb6bc8298
class ReporteView(ListView):
    model = OrdenDespacho
    template_name = 'reportebackend.html'
    context_object_name = 'ordenes_despacho'
    paginate_by = 5  # Número de objetos por página

    def get_queryset(self):
        # Obtiene el valor de búsqueda y las fechas desde la URL (si existe)
        search_query = self.request.GET.get('search', '')
        fecha_inicio = self.request.GET.get('fecha_inicio')
        fecha_fin = self.request.GET.get('fecha_fin')

        queryset = OrdenDespacho.objects.select_related('cliente', 'cliente__cliente_profile','transportista')

        # Filtro de búsqueda por nombre de supermercado (insensible a mayúsculas/minúsculas)
        if search_query:
            queryset = queryset.filter(
                Q(cliente__cliente_profile__nombre_supermercado__icontains=search_query)
            )

        if fecha_inicio:
            fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d')
            queryset = queryset.filter(fecha_creacion__gte=fecha_inicio)  # Filtramos por fecha_creacion
        if fecha_fin:
            fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d')
            queryset = queryset.filter(fecha_creacion__lte=fecha_fin)  # Filtramos por fecha_creacion

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Añadimos el valor de búsqueda para que se mantenga en el campo del formulario
        context['search'] = self.request.GET.get('search', '')
        context['fecha_inicio'] = self.request.GET.get('fecha_inicio', '')
        context['fecha_fin'] = self.request.GET.get('fecha_fin', '')
        return context
    
def exportar_excel(request):
<<<<<<< HEAD
    # Fechas "en duro"
    fecha_inicio = '01-11-2024'
    fecha_fin = '01-12-2024'

    # Obtener las órdenes de la base de datos (manteniendo esta parte)
    queryset = OrdenDespacho.objects.select_related('cliente', 'cliente__cliente_profile', 'transportista')

=======
    
    queryset = OrdenDespacho.objects.select_related('cliente', 'cliente__cliente_profile','transportista')
>>>>>>> a78dd5d0da9c93d7f98d1c059e1ae4fdb6bc8298
    # Crear el libro de Excel
    wb = Workbook()
    ws_reporte = wb.active
    ws_reporte.title = "Reporte Financiero"

<<<<<<< HEAD
    
    # Título del reporte con fechas
    ws_reporte.append([f'REPORTE FINANCIERO - Desde {fecha_inicio} hasta {fecha_fin}'])
    ws_reporte.append([])  # Dejar una fila vacía

    # Ajustar el ancho de las columnas según los encabezados
    column_widths = [20, 40, 40, 20]  # Ancho de cada columna (puedes ajustar estos valores)
    for i, column_width in enumerate(column_widths, start=1):
        col_letter = chr(64 + i)  # Convertir índice numérico a letra (1 -> A, 2 -> B, 3 -> C, 4 -> D)
        ws_reporte.column_dimensions[col_letter].width = column_width

    # Encabezados de la tabla
    headers = ['ID de la Orden', 'Cliente', 'Asignado a', 'Total de la Orden']
    ws_reporte.append(headers)

    # Crear bordes para las celdas
    thin_border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))

    # Sumar el total de las órdenes (simularemos el total)
    total_ordenes = 0

    # Agregar los datos de las órdenes al Excel
    for orden in queryset:
        cliente_nombre = orden.cliente.cliente_profile.nombre_supermercado if orden.cliente and orden.cliente.cliente_profile else "N/A"
        transportista_nombre = orden.transportista.get_full_name() if orden.transportista else "N/A"

        # Simular el total de la orden con valores aleatorios (diferentes para cada orden)
        total_orden_simulado = random.randint(200000, 500000)  # Generar un valor aleatorio entre 200.000 y 500.000 CLP

        # Añadir la fila con los datos y el total simulado
        row = [str(orden.id), cliente_nombre, transportista_nombre, total_orden_simulado]
        ws_reporte.append(row)
        total_ordenes += total_orden_simulado  # Sumar el total simulado

    # Aplicar bordes a cada celda de la fila de datos
    for row in ws_reporte.iter_rows(min_row=2, max_row=ws_reporte.max_row):
        for cell in row:
            cell.border = thin_border

    # Agregar una fila para el total final
    ws_reporte.append([])  # Espacio vacío
    ws_reporte.append(["", "", "Total Final", total_ordenes])  # Mostrar el total final simulado

    # Definir el rango de la tabla (por ejemplo, desde la fila 2 hasta la fila con los datos)
    tabla_range = f'A2:D{len(queryset) + 3}'  # Incluir las filas de los datos y el total

    # Crear la tabla
    table = Table(displayName="ReporteTabla", ref=tabla_range)

    # Estilos para la tabla (bordes en blanco y negro, sin colores)
    style = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=False, showColumnStripes=False)
    table.tableStyleInfo = style

    # Agregar la tabla al worksheet
    ws_reporte.add_table(table)

    # Dar formato a las celdas de la primera fila (encabezados)
    for cell in ws_reporte[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Ajustar el ancho de las columnas
    column_widths = [20, 40, 40, 20]  # Ajustar el tamaño de las columnas según el contenido
    for i, column_width in enumerate(column_widths, start=1):
        col_letter = chr(64 + i)  # Convertir índice numérico a letra (1 -> A, 2 -> B, 3 -> C, 4 -> D)
        ws_reporte.column_dimensions[col_letter].width = column_width

    # Crear la respuesta HTTP con el archivo Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    # El nombre del archivo
    response['Content-Disposition'] = 'attachment; filename=Reporte_financiero.xlsx'

    try:
        # Guardar el archivo Excel en la respuesta HTTP
        wb.save(response)
    except Exception as e:
        return HttpResponse(f"Hubo un problema con la generación del archivo Excel: {str(e)}")

    return response


def exportar_pdf(request):
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    if not fecha_inicio or not fecha_fin:
        return HttpResponse("Error: Las fechas de inicio y fin son necesarias.", status=400)

    try:
        fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d')
        fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d')
    except ValueError:
        return HttpResponse("Error: El formato de las fechas es incorrecto.", status=400)

    # Consulta a la base de datos
    queryset = OrdenDespacho.objects.select_related('cliente', 'cliente__cliente_profile', 'transportista')

    ordenes_data = []_
    total_general = 0

    if queryset.exists():
        # Si hay órdenes en la base de datos
        for orden in queryset:
            cliente_nombre = orden.cliente.cliente_profile.nombre_supermercado if orden.cliente and orden.cliente.cliente_profile else "N/A"
            transportista_nombre = orden.transportista.get_full_name() if orden.transportista else "N/A"
            
            # Simulamos el total de la orden con un número aleatorio entre 200.000 y 500.000 CLP
            total_orden_simulado = random.randint(200000, 500000)

            ordenes_data.append({
                'id': str(orden.id),
                'cliente': cliente_nombre,
                'asignado_a': transportista_nombre,
                'total_orden': total_orden_simulado
            })
            total_general += total_orden_simulado  # Acumulamos el total general
    else:
        # Si no hay órdenes, generamos datos aleatorios para 10 órdenes
        min_precio = 5000
        max_precio = 100000
        for i in range(10):  # Generamos 10 órdenes aleatorias
            total_orden_simulado = random.randint(min_precio, max_precio)  # Total aleatorio
            total_general += total_orden_simulado

            ordenes_data.append({
                'id': str(i + 1),  # ID de la orden
                'cliente': f'Cliente {i + 1}',  # Nombre de cliente simulado
                'asignado_a': f'Empleado {random.randint(1, 5)}',  # Asignado a un empleado aleatorio
                'total_orden': total_orden_simulado  # Total aleatorio de la orden
            })

    # Renderizar el HTML para el PDF
    html_content = render_to_string('reporte_pdf_template.html', {
        'ordenes': ordenes_data,
        'fecha_inicio': fecha_inicio.strftime('%d-%m-%Y'),
        'fecha_fin': fecha_fin.strftime('%d-%m-%Y'),
        'total_general': total_general
    })

    # Generación del PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_ordenes.pdf"'

    pisa_status = pisa.CreatePDF(html_content, dest=response)

    if pisa_status.err:
        return HttpResponse('Error al generar el PDF', status=400)

    return response


#def exportar_pdf(request):
    # Obtener las fechas de inicio y fin desde los parámetros de la solicitud
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    # Asegurarnos de que las fechas están en el formato correcto (si no están vacías)
    if fecha_inicio and fecha_fin:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d')  # Convertir de string a datetime
            fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d')  # Convertir de string a datetime
        except ValueError:
            return HttpResponse('Fecha inválida', status=400)
    else:
        # Si no se proporcionan fechas, usamos el rango completo (sin filtro de fecha)
        fecha_inicio = None
        fecha_fin = None

    # Obtén los datos de las órdenes de despacho, aplicando el filtro de fechas si es necesario
    queryset = OrdenDespacho.objects.select_related('cliente', 'cliente__cliente_profile', 'transportista')

    # Filtrar por fechas si se proporcionaron
    if fecha_inicio and fecha_fin:
        queryset = queryset.filter(fecha_creacion__range=[fecha_inicio, fecha_fin])

=======
    # Agregar parámetros del reporte
    ws_reporte.append(['REPORTE FINANCIERO'])
    ws_reporte.append(['Id de la orden', 'Cliente', 'Asignado a'])

    for orden in queryset:
        # Añadir una fila con los datos correspondientes
        cliente_nombre = orden.cliente.cliente_profile.nombre_supermercado if orden.cliente and orden.cliente.cliente_profile else "N/A"
        transportista_nombre = orden.transportista.get_full_name() if orden.transportista else "N/A"
        
        ws_reporte.append([str(orden.id), cliente_nombre, transportista_nombre])


    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Reporte_financiero_{}.xlsx'.format(
        timezone.now().strftime('%Y%m%d_%H%M%S')
    )

    wb.save(response)
    return response
    #return HttpResponse("¡La función exportar_excel ha sido ejecutada!")


def exportar_pdf(request):
    # Obtén los datos de las órdenes de despacho
    queryset = OrdenDespacho.objects.select_related('cliente', 'cliente__cliente_profile', 'transportista')

>>>>>>> a78dd5d0da9c93d7f98d1c059e1ae4fdb6bc8298
    # Convertimos los UUIDs a cadenas (si es necesario)
    ordenes_data = []
    for orden in queryset:
        orden_id_str = str(orden.id)  # Convertimos el UUID a string
        cliente_nombre = orden.cliente.cliente_profile.nombre_supermercado
        asignado_a = orden.transportista.get_full_name() if orden.transportista else 'N/A'
        
        ordenes_data.append({
            'id': orden_id_str,
            'cliente': cliente_nombre,
<<<<<<< HEAD
            'asignado_a': asignado_a,
            'fecha_creacion': orden.fecha_creacion.strftime('%Y-%m-%d')  # Formatear la fecha
        })

    # Renderizamos el HTML a partir de una plantilla
    html_content = render_to_string('reporte_pdf_template.html', {
        'ordenes': ordenes_data,
        'fecha_inicio': fecha_inicio.strftime('%Y-%m-%d') if fecha_inicio else 'N/A',
        'fecha_fin': fecha_fin.strftime('%Y-%m-%d') if fecha_fin else 'N/A'
    })
=======
            'asignado_a': asignado_a
        })

    # Renderizamos el HTML a partir de una plantilla
    html_content = render_to_string('reporte_pdf_template.html', {'ordenes': ordenes_data})
>>>>>>> a78dd5d0da9c93d7f98d1c059e1ae4fdb6bc8298

    # Creamos una respuesta HTTP con el contenido PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_ordenes.pdf"'

    # Usamos xhtml2pdf para generar el PDF a partir del HTML
    pisa_status = pisa.CreatePDF(html_content, dest=response)

    # Si hubo un error al generar el PDF, retornamos un mensaje
    if pisa_status.err:
        return HttpResponse('Error al generar el PDF', status=400)

    return response