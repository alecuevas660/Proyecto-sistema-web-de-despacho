from uuid import UUID
import openpyxl
import locale
from decimal import Decimal
from django.shortcuts import get_object_or_404, redirect, render
from apps.ventas.models import Venta
from apps.inventario.models import Categoria, OrdenDespacho, Product, SeguimientoEnvio
from django.utils import timezone
from datetime import datetime, timedelta
from xhtml2pdf import pisa
from io import BytesIO
from django.template.loader import render_to_string
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl import Workbook
from django.http import JsonResponse, HttpResponseRedirect, HttpResponse
from django.template.loader import get_template
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, PageBreak, Spacer, Paragraph
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

def lista_envios(request):
    """Vista que muestra una lista de todos los envíos con botones para generar los informes."""
    envios = SeguimientoEnvio.objects.all()

    # Revisar si se ha solicitado descargar el PDF o el Excel
    if 'descargar_pdf' in request.GET:
        envio_id = request.GET.get('envio_id')
        return informe_envio_pdf(request, envio_id)

    if 'descargar_excel' in request.GET:
        envio_id = request.GET.get('envio_id')
        return informe_envio_excel(request, envio_id)

    return render(request, 'reportes/lista_envios.html', {'envios': envios})
	
def lista_ventas(request):
    """Vista que muestra una lista de todas las ventas con botones para generar los informes."""
    ventas = Venta.objects.all()
    
    # Revisar si se ha solicitado descargar el PDF o el Excel
    if 'descargar_pdf' in request.GET:
        venta_id = request.GET.get('venta_id')
        return informe_venta_pdf(request, venta_id)

    if 'descargar_excel' in request.GET:
        venta_id = request.GET.get('venta_id')
        return informe_venta_excel(request, venta_id)

    return render(request, 'reportes/lista_ventas.html', {'ventas': ventas})
	
def lista_inventario(request):
    # Obtener filtros del request
    nombre = request.GET.get('nombre', '')
    categoria_id = request.GET.get('categoria')
    stock_status = request.GET.get('stock_status')
    fecha_filtro = request.GET.get('fecha_filtro')

    # Filtrar productos activos
    productos = Product.objects.filter(activo=True)

    # Filtro de nombre
    if nombre:
        productos = productos.filter(name__icontains=nombre)

    # Filtro de categoría
    if categoria_id:
        productos = productos.filter(categoria__id=categoria_id)

    # Filtro de estado de stock
    if stock_status:
        productos = [p for p in productos if p.get_stock_status()[1] == stock_status]

    # Filtro de fecha
    if fecha_filtro == 'hoy':
        productos = productos.filter(created_at__date=timezone.now().date())
    elif fecha_filtro == 'ultimo_mes':
        productos = productos.filter(created_at__gte=timezone.now() - timedelta(days=30))
    elif fecha_filtro == 'ultimo_ano':
        productos = productos.filter(created_at__gte=timezone.now() - timedelta(days=365))

    # Crear lista de productos con stock y valores calculados
    productos_inventario = []
    total_valor_inventario = 0

    for producto in productos:
        stock_actual = producto.get_stock_actual()
        valor_inventario = producto.price * stock_actual
        total_valor_inventario += valor_inventario

        productos_inventario.append({
            'producto': producto,
            'stock_actual': stock_actual,
            'valor_inventario': valor_inventario,
            'stock_status': producto.get_stock_status()
        })

    # Pasar datos y categorías para los filtros al template
    categorias = Categoria.objects.all()

    # Verificar si se solicita descargar el PDF
    if 'descargar_pdf' in request.GET:
        return generar_pdf_reporte_inventario(request, productos_inventario, total_valor_inventario, categorias)
    
    # Verificar si se solicita descargar el Excel
    if 'descargar_excel' in request.GET:
        return generar_excel_reporte_inventario(productos_inventario, total_valor_inventario)

    # Renderizar el reporte en la vista HTML normal
    return render(request, 'reportes/reporte_inventario.html', {
        'productos_inventario': productos_inventario,
        'total_valor_inventario': Decimal(total_valor_inventario),
        'categorias': categorias,
    })

def informe_envio_pdf(request, envio_id):
    """Genera un informe de seguimiento de envío en PDF."""
    # Obtener el objeto de SeguimientoEnvio, lanzar 404 si no existe
    envio = get_object_or_404(SeguimientoEnvio, pk=envio_id)

    # Asegurarse de que el envío tiene una orden asociada
    orden_despacho = envio.orden
    if not orden_despacho:
        return HttpResponse('No se encuentra la orden de despacho asociada al seguimiento.', status=404)

    # Crear los datos para el informe
    data = {
        'orden_id': orden_despacho.id,
        'cliente': orden_despacho.cliente.get_full_name(),
        'transportista': orden_despacho.transportista.get_full_name(),
        'direccion_entrega': orden_despacho.direccion_entrega,
        'estado_envio': envio.get_estado_envio_display(),
        'ubicacion_actual': envio.ubicacion_actual,
        'comentarios': envio.comentarios,
        'fecha_actualizacion': envio.fecha_actualizacion,
    }

    # Renderizar la plantilla HTML para el PDF
    template = get_template('reportes/informe_envio_pdf.html')
    html = template.render(data)

    # Generar el archivo PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="informe_envio_{orden_despacho.id}.pdf"'

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error al generar el PDF', status=500)

    return response

def informe_envio_excel(request, envio_id):
    """Genera un informe de seguimiento de envío en Excel."""
    # Obtener el objeto de SeguimientoEnvio, lanzar 404 si no existe
    envio = get_object_or_404(SeguimientoEnvio, pk=envio_id)

    # Asegurarse de que el envío tiene una orden asociada
    orden_despacho = envio.orden
    if not orden_despacho:
        return HttpResponse('No se encuentra la orden de despacho asociada al seguimiento.', status=404)

    # Crear el libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Seguimiento Orden {orden_despacho.id}"

    # Encabezados
    headers = [
        'ID de la Orden',
        'Cliente',
        'Transportista',
        'Dirección de Entrega',
        'Estado de Envío',
        'Ubicación Actual',
        'Última Actualización',
        'Comentarios'
    ]
    ws.append(headers)

    # Agregar los datos del seguimiento de envío
    ws.append([
        str(orden_despacho.id),
        orden_despacho.cliente.get_full_name(),
        orden_despacho.transportista.get_full_name(),
        orden_despacho.direccion_entrega,
        envio.get_estado_envio_display(),
        envio.ubicacion_actual,
        envio.fecha_actualizacion.strftime('%Y-%m-%d %H:%M:%S'),
        envio.comentarios,
    ])

    # Ajustar el ancho de las columnas para mejor legibilidad
    for col in ws.columns:
        max_length = max(len(str(cell.value) if cell.value else '') for cell in col)
        col_letter = col[0].column_letter  # Obtener la letra de la columna
        ws.column_dimensions[col_letter].width = max_length + 2  # Ajustar ancho

    # Preparar la respuesta para descarga
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="informe_envio_{orden_despacho.id}.xlsx"'

    # Guardar el archivo en la respuesta
    wb.save(response)
    return response

def informe_venta_pdf(request, venta_id):
    """Genera un informe de venta específico en PDF."""
    venta = get_object_or_404(Venta, pk=venta_id)

    # Crear los datos para el informe
    data = {
        'cliente': venta.cliente.cliente_profile.nombre_supermercado,
        'asignado_despacho': venta.vendedor.get_full_name(),
        'productos': [
            {
                'nombre': detalle.producto.name,
                'cantidad': detalle.cantidad,
                'precio_unitario': detalle.precio_unitario,
                'subtotal': detalle.subtotal,
            }
            for detalle in venta.detalles.all()
        ],
        'total': venta.total,
    }

    # Renderizar la plantilla HTML para el PDF
    template = get_template('reportes/informe_venta_pdf.html')
    html = template.render(data)

    # Generar el archivo PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="informe_venta_{venta.numero_venta}.pdf"'

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error al generar el PDF', status=500)

    return response

def informe_venta_excel(request, venta_id):
    """Genera un informe de venta específico en Excel."""
    venta = get_object_or_404(Venta, pk=venta_id)

    # Crear el libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Venta {venta.numero_venta}"

    # Encabezados
    headers = ['Cliente', 'Asignado Despacho', 'Producto', 'Cantidad', 'Precio Unitario', 'Subtotal', 'Total']
    ws.append(headers)

    # Agregar los detalles de la venta
    for detalle in venta.detalles.all():
        ws.append([
            venta.cliente.cliente_profile.nombre_supermercado,
            venta.vendedor.get_full_name(),
            detalle.producto.name,
            detalle.cantidad,
            detalle.precio_unitario,
            detalle.subtotal,
            venta.total,
        ])
    
    # Ajustar el ancho de las columnas para mejor legibilidad
    for col in ws.columns:
        max_length = max(len(str(cell.value) if cell.value else '') for cell in col)
        col_letter = col[0].column_letter  # Obtener la letra de la columna
        ws.column_dimensions[col_letter].width = max_length + 2  # Ajustar ancho

    # Preparar la respuesta para descarga
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="informe_venta_{venta.numero_venta}.xlsx"'

    wb.save(response)
    return response

# Establecer el locale para utilizar el punto como separador de miles
locale.setlocale(locale.LC_ALL, 'es_CL.UTF-8')  # Usamos la configuración de Chile para que use el punto como separador

def eliminar_zeros(cantidad):
    """Elimina los ceros innecesarios y usa el punto como separador de miles."""
    # Formatear el número con el separador de miles
    cantidad_formateada = locale.format_string('%.2f', cantidad, grouping=True)
    
    # Eliminar los ceros innecesarios si es posible
    if cantidad_formateada.endswith(",00"):
        return cantidad_formateada[:-3]  # Eliminar el ".00"
    return cantidad_formateada

def generar_pdf_reporte_inventario(request, productos_inventario, total_valor_inventario, categorias):
    """Genera el PDF del reporte inventario en ReportLab."""

    # Crear un buffer para almacenar el PDF
    buffer = BytesIO()
    
    # Crear el documento PDF
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    
    styles = getSampleStyleSheet()

    # Estilo personalizado para el título
    title_style = ParagraphStyle(
        name="TitleStyle",
        fontName="Helvetica-Bold",
        fontSize=16,
        textColor=colors.black,
        alignment=1,  # Centrado
        spaceAfter=12  # Espacio después del título
    )

    # Título del reporte
    title = Paragraph("Reporte de Inventario", title_style)

    # Crear una lista para las filas de la tabla
    data = []
    
    # Agregar encabezados
    headers = ['Nombre', 'Categoría', 'Descripción', 'Precio', 'Stock Mínimo', 'Stock Actual', 'Estado', 'Valor Total en Inventario']
    data.append(headers)

    # Agregar los productos a la tabla
    for item in productos_inventario:
        producto = item['producto']
        stock_status = item['stock_status'][0]
        valor_inventario = item['valor_inventario']

        # Asegurarse de que la descripción no se salga del PDF, dividiendo las palabras largas
        descripcion = producto.description
        if len(descripcion) > 22:  # Ajusta este límite según sea necesario
            descripcion = '\n'.join([descripcion[i:i+22] for i in range(0, len(descripcion), 22)])
        
        nombre = producto.description
        if len(nombre) > 14:  # Ajusta este límite según sea necesario
            nombre = '\n'.join([nombre[i:i+14] for i in range(0, len(nombre), 14)])

        # Usar la función para eliminar los ceros innecesarios y con el formato correcto
        precio_formateado = eliminar_zeros(producto.price)
        valor_inventario_formateado = eliminar_zeros(valor_inventario)
        
        row = [
            nombre,
            producto.categoria.nombre,
            descripcion,  # La descripción ya cortada para que no se salga
            f"${precio_formateado}",
            producto.stock_minimo,
            item['stock_actual'],
            stock_status,
            f"${valor_inventario_formateado}"
        ]
        data.append(row)

    # Definir el ancho de las columnas (reducido aún más para que todo quepa)
    col_widths = [1.0 * inch, 1.2 * inch, 1.5 * inch, 0.8 * inch, 0.8 * inch, 0.8 * inch, 1.0 * inch, 1.0 * inch]

    # Crear la tabla con los datos
    table = Table(data, colWidths=col_widths)

    # Estilos de la tabla (manejo de palabras largas, bordes, etc.)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  # Encabezado gris
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('LINEBEFORE', (0, 0), (0, -1), 0.5, colors.black),  # Línea antes de cada celda
        ('FONTSIZE', (0, 0), (-1, -1), 8),  # Fuente pequeña para mejor ajuste
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),  # Color de texto
        ('WORDSPACE', (0, 1), (-1, -1), 5),  # Para permitir espacio entre palabras
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),  # Alineación izquierda en todas las celdas
        ('CANVASSPLIT', (0, 0), (-1, -1), True),  # Permitir el ajuste de texto dentro de las celdas
    ]))

    # Crear una lista de elementos que se agregarán al documento
    elements = [title, Spacer(1, 12), table]  # Título seguido de un espacio y la tabla
    
    # Si el contenido es muy largo, agregar un salto de página
    if len(productos_inventario) > 20:  # Ajusta este número según lo que caben en una página
        elements.append(PageBreak())

    # Agregar el total al final
    elements.append(Spacer(1, 12))  # Espacio antes del total
    elements.append(Paragraph(f"Valor Total del Inventario: ${eliminar_zeros(total_valor_inventario)}", styles['Normal']))

    # Crear el documento PDF
    doc.build(elements)

    # Configurar la respuesta para que sea un archivo PDF
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_inventario.pdf"'

    return response

def generar_excel_reporte_inventario(productos_inventario, total_valor_inventario):
    # Crear un nuevo Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Inventario"

    # Agregar encabezados
    headers = [
        "Nombre del Producto", 
        "Categoría", 
        "Descripción", 
        "Precio", 
        "Stock Mínimo", 
        "Stock Actual", 
        "Estado del Stock", 
        "Valor Total en Inventario"
    ]
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    # Agregar datos de productos
    for row_num, producto in enumerate(productos_inventario, start=2):
        ws.cell(row=row_num, column=1, value=producto['producto'].name)
        ws.cell(row=row_num, column=2, value=producto['producto'].categoria.nombre)
        ws.cell(row=row_num, column=3, value=producto['producto'].description)
        ws.cell(row=row_num, column=4, value=float(producto['producto'].price))
        ws.cell(row=row_num, column=5, value=producto['producto'].stock_minimo)
        ws.cell(row=row_num, column=6, value=producto['stock_actual'])
        ws.cell(row=row_num, column=7, value=producto['stock_status'][0])  # Estado
        ws.cell(row=row_num, column=8, value=float(producto['valor_inventario']))

    # Agregar total del inventario al final
    total_row = len(productos_inventario) + 2
    ws.cell(row=total_row, column=1, value="Valor Total del Inventario")
    ws.cell(row=total_row, column=8, value=float(total_valor_inventario))
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    
    # Ajustar el ancho de las columnas para mejor legibilidad
    for col in ws.columns:
        max_length = max(len(str(cell.value) if cell.value else '') for cell in col)
        col_letter = col[0].column_letter  # Obtener la letra de la columna
        ws.column_dimensions[col_letter].width = max_length + 2  # Ajustar ancho

    # Preparar respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename="reporte_inventario.xlsx"'

    # Guardar el Workbook en el response
    wb.save(response)
    return response

def reporte_general_inventario(): #Plantilla para realizar reporte general
    productos = Product.objects.filter(activo=True)
    reporte_data = []
    total_valor_inventario = Decimal('0.00')

    for producto in productos:
        stock_actual = producto.get_stock_actual()
        stock_status, stock_status_class = producto.get_stock_status()
        valor_total_producto = producto.price * stock_actual

        reporte_data.append({
            'nombre': producto.name,
            'categoria': producto.categoria.name,
            'precio': producto.price,
            'stock_minimo': producto.stock_minimo,
            'stock_actual': stock_actual,
            'estado_stock': stock_status,
            'valor_total': valor_total_producto,
        })

        total_valor_inventario += valor_total_producto

    return reporte_data, total_valor_inventario

def reporte_por_categoria_inventario(categoria_id): #Plantilla para realizar reporte por categoria
    categoria = Categoria.objects.get(id=categoria_id)
    productos = categoria.productos.filter(activo=True)
    reporte_data = []
    total_valor_inventario = Decimal('0.00')

    for producto in productos:
        stock_actual = producto.get_stock_actual()
        stock_status, stock_status_class = producto.get_stock_status()
        valor_total_producto = producto.price * stock_actual

        reporte_data.append({
            'nombre': producto.name,
            'descripcion': producto.description,
            'precio': producto.price,
            'stock_minimo': producto.stock_minimo,
            'stock_actual': stock_actual,
            'estado_stock': stock_status,
            'valor_total': valor_total_producto,
        })

        total_valor_inventario += valor_total_producto

    return reporte_data, total_valor_inventario

def reporte_bajos_en_stock_inventario(): #Plantilla para realizar reporte bajos en stock
    productos = Product.objects.filter(activo=True)
    reporte_data = []
    total_valor_inventario = Decimal('0.00')

    for producto in productos:
        stock_actual = producto.get_stock_actual()
        if stock_actual < producto.stock_minimo * 0.30:  # Si el stock es menor al 30% del stock mínimo
            stock_status, stock_status_class = producto.get_stock_status()
            valor_total_producto = producto.price * stock_actual

            reporte_data.append({
                'nombre': producto.name,
                'categoria': producto.categoria.name,
                'precio': producto.price,
                'stock_minimo': producto.stock_minimo,
                'stock_actual': stock_actual,
                'estado_stock': stock_status,
                'valor_total': valor_total_producto,
            })

            total_valor_inventario += valor_total_producto

    return reporte_data, total_valor_inventario

from django.core.paginator import Paginator
from django.shortcuts import render
from datetime import datetime, timedelta
from .models import Reporte

def reporte_view(request):
    # Obtener el filtro de fecha desde la URL (por defecto es 'hoy')
    filtro = request.GET.get('filtro', 'hoy')
    
    # Obtener el término de búsqueda desde la URL (por defecto es una cadena vacía)
    query = request.GET.get('q', '')  # 'q' es el nombre del parámetro de búsqueda
    
    # Obtener la fecha actual
    hoy = datetime.now()
    
    # Filtrar según el criterio de fecha
    if filtro == 'hoy':
        fecha_inicio = hoy.replace(hour=0, minute=0, second=0, microsecond=0)  # Comienza a las 00:00 de hoy
        fecha_fin = hoy.replace(hour=23, minute=59, second=59, microsecond=999999)  # Termina a las 23:59 de hoy
        reportes = Reporte.objects.filter(fecha__range=[fecha_inicio, fecha_fin])
        
    elif filtro == 'mes':
        fecha_inicio = hoy - timedelta(days=hoy.day)  # Fecha del primer día del mes
        fecha_inicio = fecha_inicio.replace(hour=0, minute=0, second=0, microsecond=0)
        reportes = Reporte.objects.filter(fecha__gte=fecha_inicio)
        
    elif filtro == 'año':
        fecha_inicio = hoy.replace(month=1, day=1)  # Primer día del año
        fecha_inicio = fecha_inicio.replace(hour=0, minute=0, second=0, microsecond=0)
        reportes = Reporte.objects.filter(fecha__gte=fecha_inicio)
    
    # Si hay una búsqueda, filtrar también por el nombre o descripción
    if query:
        reportes = reportes.filter(nombre__icontains=query)  # Filtra los reportes por nombre que contengan 'query'
    
    # Paginación: dividir los reportes en páginas de 10 elementos
    paginator = Paginator(reportes, 10)  # 10 reportes por página
    
    # Obtener el número de la página actual
    page_number = request.GET.get('page')  # Esto puede ser pasado en la URL, como ?page=2
    page_obj = paginator.get_page(page_number)  # Obtener la página actual
    
    return render(request, 'reportes.html', {
        'page_obj': page_obj,
        'filtro': filtro,
        'query': query  # Pasamos el término de búsqueda al template
    })

