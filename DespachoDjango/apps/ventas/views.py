import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, render
from DespachoDjango.apps.ventas.models import Venta, DetalleVenta

# Create your views here.
def reporte_venta(request, venta_id):
    # Obtener la venta y sus detalles
    venta = get_object_or_404(Venta, id=venta_id)
    detalles = DetalleVenta.objects.filter(venta=venta)  # Usar DetalleVenta para obtener los detalles de la venta

    # Calcular el total de la venta
    total_venta = venta.total

    # Contexto para el template
    context = {
        'cliente': venta.cliente.get_full_name(),
        'vendedor': venta.vendedor.get_full_name(),
        'detalles': detalles,
        'total': total_venta,
    }

    return render(request, 'reporte_venta.html', context)

def reporte_venta_excel(request, venta_id):
    # Obtener la venta y sus detalles
    venta = get_object_or_404(Venta, id=venta_id)
    detalles = DetalleVenta.objects.filter(venta=venta)

    # Crear el archivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Reporte de Venta {venta.numero_venta}"

    # Encabezado
    ws['A1'] = "Reporte de Venta"
    ws['A2'] = f"Número de Venta: {venta.numero_venta}"
    ws['A3'] = f"Cliente: {venta.cliente.get_full_name()}"
    ws['A4'] = f"Asignado de Despacho: {venta.vendedor.get_full_name()}"
    ws['A1'].font = Font(size=14, bold=True)

    # Encabezados de la tabla de productos
    headers = ["Producto", "Cantidad", "Precio Unitario", "Subtotal"]
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=6, column=col_num, value=header).font = Font(bold=True)

    # Agregar los detalles de la venta
    row_num = 7
    for detalle in detalles:
        ws.cell(row=row_num, column=1, value=detalle.producto.name)
        ws.cell(row=row_num, column=2, value=detalle.cantidad)
        ws.cell(row=row_num, column=3, value=float(detalle.precio_unitario))
        ws.cell(row=row_num, column=4, value=float(detalle.subtotal))
        row_num += 1

    # Total de la venta
    ws.cell(row=row_num, column=3, value="Total:")
    ws.cell(row=row_num, column=4, value=float(venta.total)).font = Font(bold=True)

    # Preparar la respuesta HTTP para descargar el archivo
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = f'attachment; filename="reporte_venta_{venta.numero_venta}.xlsx"'

    # Guardar el archivo Excel en la respuesta
    wb.save(response)
    return response